import os
import re
import shutil
from dotenv import load_dotenv
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_openai import ChatOpenAI
from langchain_groq import ChatGroq
from langchain_core.messages import SystemMessage, HumanMessage
import chainlit as cl
from google.api_core import exceptions as google_exceptions
from google.api_core.exceptions import ServiceUnavailable, InternalServerError
import pandas as pd
import numpy as np
import pymysql
import pymysql.cursors
import openpyxl
from tabulate import tabulate

# .env 파일 로드`
load_dotenv()

# 실제 분석에 사용할 메인 모델
MAIN_MODEL_NAME = "gemini-3-flash-preview"

# API Key 검증용 가벼운 모델 (속도 우선)
TEST_MODEL_NAME = "gemini-2.5-flash-lite"

# LLM 모델 설정
#llm = ChatGoogleGenerativeAI(model="gemini-2.5-flash", temperature=0)
#llm = ChatGroq(model_name="openai/gpt-oss-120b", temperature=0)
# llm = ChatOpenAI(
#     api_key=OPENROUTER_API_KEY,
#     base_url="https://openrouter.ai/api/v1",
#     model="amazon/nova-2-lite-v1:free", #google/gemini-2.0-flash-exp:free
#     temperature=0,
#     default_headers={
#         "HTTP-Referer": getenv("kungmo2.mooo.com:8501"), # Optional. Site URL for rankings on openrouter.ai.
#         "X-Title": getenv("학교생기부 점검 도우미"), # Optional. Site title for rankings on openrouter.ai.
#     }
# )

# --- [수정] DB 로깅 함수 (디버깅 강화) ---
def log_to_db(session_id, subject, success, error_message=None):
    """
    MariaDB에 사용 이력을 저장합니다.
    에러 발생 시 콘솔에 원인을 출력합니다.
    """
    conn = None
    try:
        # DB 연결 정보 확인
        host = os.getenv("DB_HOST", "localhost")
        user = os.getenv("DB_USER", "root")
        password = os.getenv("DB_PASSWORD", "")
        db_name = os.getenv("DB_NAME", "school_records")
        port = int(os.getenv("DB_PORT"))

        conn = pymysql.connect(
            host=host,
            user=user,
            password=password,
            database=db_name,
            port=port,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=5 # 연결 타임아웃 설정
        )
        
        with conn.cursor() as cursor:
            sql = """
            INSERT INTO school_record_logs (session_id, subject, success, error_message, created_at)
            VALUES (%s, %s, %s, %s, NOW())
            """
            # success는 boolean -> int(1/0) 변환
            # error_message가 너무 길 경우를 대비해 짤라서 저장 (TEXT 타입 한계 고려, 보통 충분함)
            safe_error_msg = str(error_message)[:2000] if error_message else None
            
            cursor.execute(sql, (session_id, subject, int(success), safe_error_msg))
        
        conn.commit()
        print(f"[DB Success] 로그 저장 완료: {subject}, Success={success}")

    except pymysql.MySQLError as e:
        # DB 관련 에러는 여기서 잡힙니다. 터미널 로그를 확인해주세요.
        print(f"============== [DB Error] ==============")
        print(f"Error Code: {e.args[0]}")
        print(f"Error Message: {e.args[1]}")
        print(f"DB Config: Host={host}, User={user}, DB={db_name}")
        print(f"========================================")
    except Exception as e:
        print(f"[DB Log Unexpected Error] {e}")
    finally:
        if conn:
            conn.close()

async def validate_api_key(api_key: str):
    """
    입력받은 API Key가 유효한지 'gemini-2.5-flash-lite'로 빠르게 확인합니다.
    """
    try:
        # 검증용 Lite 모델 생성
        test_llm = ChatGoogleGenerativeAI(model=TEST_MODEL_NAME, google_api_key=api_key, temperature=0)
        # 1토큰 정도의 매우 짧은 응답을 유도하여 연결 확인
        await test_llm.ainvoke([HumanMessage(content="Hi")])
        return True
    except Exception as e:
        print(f"Google API 키가 유효하지 않습니다: {e}")
        return False

# --- 과목별 시스템 프롬프트 정의 ---
PROMPTS = {
    "과목별 세부능력 및 특기사항 (생기부 영역별 출력)": """
# Role: 고등학교 과목별 세부능력 및 특기사항 내용을 꼼꼼히 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# 생활기록부 본문 내용 parsing 상태: 과목과 학생 번호(No.) 학년, 학기를 기준으로 동료 교사가 쓴 세부능력 및 특기사항이 쓰여 있다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 학생이 '대회'에 참석했거나 특정 '강사'의 활동에 참여했다거나 '논문'을 썼다는 내용이 있는지 점검한다.
    5. 불필요하게 영어 및 외국어로 서술했는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 번호: 1, 학생 이름: (학생명), 과목: (검토한 과목), 학년/학기: (O학년 O학기)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 번호: 2, 학생 이름: (학생명), 과목: (검토한 과목), 학년/학기: (O학년 O학기)
...

## [종합 검토 의견]
- **주요 수정 사항 요약**: (반드시 고쳐야 할 부분 리스트)""",
    "과세특 (교사 입력 엑셀 파일, 고등학교)": """
# Role: 고등학교 과목별 세부능력 및 특기사항 내용을 꼼꼼히 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# 생활기록부 본문 내용 parsing 상태: 교사가 쓴 세부능력 및 특기사항이 학생별로 markdown 형식으로 쓰여 있다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 학생이 '대회'에 참석했거나 특정 '강사'의 활동에 참여했다거나 '논문'을 썼다는 내용이 있는지 점검한다.
    5. 불필요하게 영어 및 외국어로 서술했는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 반/번호: (학생 반/번호), 학생 이름: (학생명), 과목: (검토한 과목)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 반/번호: (학생 반/번호), 학생 이름: (학생명), 과목: (검토한 과목)
...

## [종합 검토 의견]
- **주요 수정 사항 요약**: (반드시 고쳐야 할 부분 리스트)""",
    "과세특 (교사 입력 엑셀 파일, 중학교)": """
# Role: 중학교 과목별 세부능력 및 특기사항 내용을 꼼꼼히 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# 생활기록부 본문 내용 parsing 상태: 교사가 쓴 세부능력 및 특기사항이 학생별로 markdown 형식으로 쓰여 있다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 학생이 '대회'에 참석했거나 특정 '강사'의 활동에 참여했다거나 '논문'을 썼다는 내용이 있는지 점검한다.
    5. 불필요하게 영어 및 외국어로 서술했는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 반/번호: (학생 반/번호), 학생 이름: (학생명), 과목: (검토한 과목)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 반/번호: (학생 반/번호), 학생 이름: (학생명), 과목: (검토한 과목)
...

## [종합 검토 의견]
- **주요 수정 사항 요약**: (반드시 고쳐야 할 부분 리스트)""",
    "행동특성 및 종합의견": """
# Role: 고등학교 행동특성 및 종합의견을 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    5. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 번호: 1, 학생 이름: (학생명), 학년: (O학년)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 번호: 2, 학생 이름: (학생명), 학년: (O학년)
...

## [종합 검토 의견]
- **주요 수정 사항 요약**: (반드시 고쳐야 할 부분 리스트)""",
    "창의적 체험활동 (고등학교 2-3학년, 현재학년)": """
# Role: 고등학교 창의적 체험활동 생활기록부를 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 동아리 활동에 쓰여 있는 두 괄호 속의 시간이 서로 다른 경우 반드시 알려줘야 한다.
    5. 특정 기관 및 상호명, 강사명, 대회 참가 내용이 쓰여 있는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 번호: 1, 학생 이름: (학생명), 학년: (O학년), 영역: (자율/동아리/진로)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 번호: 2, 학생 이름: (학생명), 학년: (O학년)""",
    "창의적 체험활동 (고등학교 2-3학년, 활동별)": """
# Role: 고등학교 창의적 체험활동 생활기록부를 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# 생활기록부 본문 내용 parsing 상태: 교사가 쓴 창의적 체험활동(자율, 동아리, 진로)의 dataframe을 markdown 형식으로 정리해 두었다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 동아리 활동에 쓰여 있는 두 괄호 속의 시간이 서로 다른 경우 반드시 알려줘야 한다.
    5. 특정 기관 및 상호명, 강사명, 대회 참가 내용이 쓰여 있는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 번호: 1, 학생 이름: (학생명), 학년: (O학년), 영역: (자율/동아리/진로)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 번호: 2, 학생 이름: (학생명), 학년: (O학년)""",
    "창의적 체험활동 (중학교, 활동별)": """
# Role: 중학교 창의적 체험활동 생활기록부를 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 동아리 활동에 쓰여 있는 이수 시간과 괄호 속에 쓰여 있는 시 간들의 합이 서로 다른 경우 반드시 알려줘야 한다.
    5. 특정 기관 및 상호명, 강사명, 대회 참가 내용이 쓰여 있는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
제공된 자료의 순서에 따라 활동별로 결과를 생성하시오.
## [(영역명)]
### 학생 번호: 1, 학생 이름: (학생명), 학년: (O학년), 영역: (자율/동아리/진로)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 번호: 2, 학생 이름: (학생명), 학년: (O학년)""",
    "창의적 체험활동 (고등학교 1학년, 현재학년)": """
# Role: 고등학교 창의적 체험활동 생활기록부를 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 동아리 활동에 쓰여 있는 두 괄호 속의 시간이 서로 다른 경우 반드시 알려줘야 한다.
    5. 특정 기관 및 상호명, 강사명, 대회 참가 내용이 쓰여 있는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 번호: 1, 학생 이름: (학생명), 학년: (O학년), 영역: (자율/동아리/진로)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)
### 학생 번호: 2, 학생 이름: (학생명), 학년: (O학년)""",
    "창의적 체험활동 (고등학교 1학년, 활동별)": """
# Role: 고등학교 창의적 체험활동 생활기록부를 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 동아리 활동에 쓰여 있는 두 괄호 속의 시간이 서로 다른 경우 반드시 알려줘야 한다.
    5. 특정 기관 및 상호명, 강사명, 대회 참가 내용이 쓰여 있는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 번호: 1, 학생 이름: (학생명), 학년: (O학년), 영역: (자율/동아리/진로)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)
### 학생 번호: 2, 학생 이름: (학생명), 학년: (O학년)""",
    "동아리 (교사 입력 엑셀 데이터 파일)": """
# Role: 고등학교 동아리 특기사항 내용을 꼼꼼히 검토하는 동료 선생님
# 시점: 현재 연도는 2025년이다.
# 주의사항: 본문 내용은 원문의 문구 그대로 사용해야만 한다.
# 생활기록부 본문 내용 parsing 상태: 교사가 쓴 동아리 특기사항의 dataframe을 tabulate하여 markdown 형식으로 쓰여 있다.
# To Dos:
    1. 오타가 있는 경우 반드시 알려줘야 한다.
    2. 문장 자체에 내재된 언어학적인 오류 여부를 점검한다.
    3. [중요] 교사의 입장에서 서술했는지 점검한다.
    4. 학생이 '대회'에 참석했거나 '논문'을 썼다는 내용이 있는지 점검한다.
    5. 특정 기관 및 상호명, 강사명, 대회 참가 내용이 쓰여 있는지 점검한다.
    6. 문장 중에 '[공백2칸]' 또는 '[공백3칸이상]' 표시가 있으면 띄어쓰기가 중복된 곳일 수 있는 확률이 있으므로 사용자에게 위치를 구체적으로 알려주면서 수정을 권한다.
    7. 문장의 종결부(마침표 직전)만 명사형 종결어미로 끝나고, 현재형으로 쓰였는지 점검한다.
       - 바른 예: ~함. ~음.
       - 틀린 예: ~한다. ~했다.
       - **주의**: 문장 중간의 연결 표현(~하며, ~하고, ~하면서 등)은 명사형 종결어미가 아니므로 점검 대상이 아니다. 오직 마침표(.) 직전의 종결어미만 점검한다.
# [중요] 답변 출력 형식:
반드시 아래의 마크다운 형식을 지켜서 답변하시오. 서론이나 잡담 없이 바로 본론으로 들어가시오.
## [학생별 정밀 검토]
### 학생 : (학생 학년/반/번호/이름)
- **검토 결과**: [적절 / 수정 필요 / 오류]
- **상세 내용**: (내용 작성)

### 학생 : (학생 학년/반/번호/이름)
...

## [종합 검토 의견]
- **주요 수정 사항 요약**: (반드시 고쳐야 할 부분 리스트)""",
}

# 공백 검출 함수
def mark_multiple_spaces(text):
    """텍스트 내의 2칸 이상 공백을 찾아 마킹합니다."""
    if not isinstance(text, str):
        return text
    # 3칸 이상 공백 우선 처리
    text = re.sub(r' {3,}', ' [공백3칸이상] ', text)
    # 2칸 공백 처리
    text = re.sub(r' {2}', ' [공백2칸] ', text)
    return text

# 과목별 세부능력 및 특기사항 변환 함수 (담임 출력)
def excel_to_clean_markdown_subject(file_path):
    """Excel 파일을 읽어서 깔끔한 Markdown으로 변환 (과목별 세부능력)"""
    try:
        # 1. CSV/Excel 파일 읽기 (헤더 없이 읽어서 구조 파악)
        df = pd.read_excel(file_path, header=None)
        
        # 2. 실제 헤더가 있는 행 찾기 ('과 목' 또는 '과목'이 포함된 행)
        header_candidates = df[df.apply(lambda row: row.astype(str).str.contains('과.*목', regex=True).any(), axis=1)]
        
        if header_candidates.empty:
            return "오류: '과목' 헤더를 찾을 수 없습니다. 파일 형식을 확인해주세요."
        
        header_row_idx = header_candidates.index[0]
        
        # 3. 헤더 설정 및 데이터 슬라이싱
        df.columns = df.iloc[header_row_idx]
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        
        # 4. 컬럼명 정리 (공백 제거)
        df.columns = [str(c).replace(' ', '').strip() for c in df.columns]
        
        # 5. 필요한 컬럼 확인 및 선택
        required_cols = ['과목', '학년', '학기', '번호', '성명', '세부능력및특기사항']
        existing_cols = [c for c in required_cols if c in df.columns]
        
        if '세부능력및특기사항' not in existing_cols:
            return "오류: '세부능력및특기사항' 컬럼을 찾을 수 없습니다."
        
        df = df[existing_cols]
        
        # 6. 불필요한 행 제거
        # 반복되는 헤더 삭제
        df = df[~df['과목'].astype(str).str.contains('과.*목', regex=True, na=False)]
        
        # 세부능력 내용이 없는 행 삭제
        df = df.dropna(subset=['세부능력및특기사항'])
        
        # [수정] 메타데이터 행 삭제 로직 개선
        # 기존: keywords 리스트에 '고등학교'가 있어 본문에 '고등학교'가 있으면 삭제되는 문제 해결
        keywords = ['학교생활기록부', '사용자명']
        
        # (1) 특정 키워드(학교생활기록부 등)가 포함된 행 찾기
        mask_keywords = df['세부능력및특기사항'].astype(str).apply(lambda x: any(k in x for k in keywords))
        
        # (2) 내용 전체가 'OO초/중/고등학교' 형식인 행 찾기 (정규표현식 사용)
        # ^:시작, $:끝 -> 문장 중간에 학교명이 있는 경우는 보존됨
        mask_school = df['세부능력및특기사항'].astype(str).str.contains(r'^\s*.*(?:초등|중|고등)학교\s*$', regex=True, na=False)
        
        # 두 조건 중 하나라도 해당되면 제거
        df = df[~(mask_keywords | mask_school)]
        
        # 7. 데이터 채우기 (Forward Fill)
        cols_to_fill = [c for c in ['과목', '학년', '학기', '번호', '성명'] if c in df.columns]
        df[cols_to_fill] = df[cols_to_fill].ffill()
        
        # 8. 번호 정리
        if '번호' in df.columns:
            df['번호'] = pd.to_numeric(df['번호'], errors='coerce')
            df = df.dropna(subset=['번호'])
            df['번호'] = df['번호'].astype(int)
        
        # 9. 그룹화하여 텍스트 합치기
        group_cols = [c for c in ['과목', '학년', '학기', '번호', '성명'] if c in df.columns]
        grouped = df.groupby(group_cols)['세부능력및특기사항'].apply(
            lambda x: mark_multiple_spaces(' '.join(x.astype(str))) # 마킹 함수 적용
        ).reset_index()
        
        # 10. 정렬
        sort_cols = [c for c in ['과목', '학기', '번호'] if c in grouped.columns]
        grouped = grouped.sort_values(by=sort_cols)
        
        # 11. Markdown 생성
        markdown_output = []
        
        for _, row in grouped.iterrows():
            subject = row.get('과목', '미상')
            name = row.get('성명', '미상')
            number = row.get('번호', '?')
            grade = row.get('학년', '?')
            semester = row.get('학기', '?')
            content = row['세부능력및특기사항']
            
            md_block = f"""
### {subject} - {name} (No. {number})
- **학년/학기:** {grade}학년 {semester}학기
- **세부능력 및 특기사항:**
> {content}

---
"""
            markdown_output.append(md_block)
        
        return "\n".join(markdown_output)
    
    except Exception as e:
        return f"오류: 파일 처리 중 문제가 발생했습니다. {str(e)}"

# 과세특 (교사 입력 엑셀 파일, 고등학교)
def xlsx_subject_to_markdown_high(file_path):
    """
    엑셀 파일을 읽어 LLM이 이해하기 쉬운 구조화된 Markdown 텍스트로 변환합니다.
    표(Table) 형식이 아닌, 학생별 섹션(Section) 형식으로 변환하여 가독성을 높입니다.
    """
    # 1. 엑셀 읽기
    # dtype=str 옵션을 주어 '반/번호'가 날짜(예: 2월 4일)로 자동 변환되는 것을 방지합니다.
    try:
        df = pd.read_excel(file_path, engine='openpyxl', dtype=str)
    except Exception as e:
        return f"파일을 읽는 중 오류가 발생했습니다: {e}"

    # 2. 필요한 컬럼 존재 여부 확인 및 추출
    required_cols = ['과목', '반/번호', '성명', '세부능력 및 특기사항']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        return f"다음 필수 컬럼이 엑셀 파일에 없습니다: {missing_cols}"
    
    df = df[required_cols].fillna('') # 결측치는 빈 문자열로 처리

    # 3. Markdown 텍스트 생성 (표 대신 문서 구조화)
    markdown_lines = []
    markdown_lines.append("# 과목별 세부능력 및 특기사항 기록")
    markdown_lines.append("---")

    for _, row in df.iterrows():
        subject = row['과목'].strip()
        name = row['성명'].strip()
        content = row['세부능력 및 특기사항'].strip()
        class_info_raw = row['반/번호'].strip()

        # 반/번호 포맷팅 (예: "2/4" -> "2반 4번"으로 변환하여 명확성 확보)
        if '/' in class_info_raw:
            parts = class_info_raw.split('/')
            if len(parts) == 2:
                class_info = f"{parts[0]}반 {parts[1]}번"
            else:
                class_info = class_info_raw
        else:
            class_info = class_info_raw

        # LLM에게 최적화된 포맷: 헤더로 학생 구분 -> 메타데이터 -> 내용
        markdown_lines.append(f"## [{subject}] {name} ({class_info})")
        markdown_lines.append(f"> **세부능력 및 특기사항**")
        markdown_lines.append(f"{content}")
        markdown_lines.append("\n---\n") # 섹션 구분선

    return "\n".join(markdown_lines)

# 과세특 (교사 입력 엑셀 파일, 중학교)
def xlsx_subject_to_markdown_mid(file_path):
    try:
        # 1. 엑셀 파일 전체 읽기 (헤더 위치를 모르므로 header=None)
        # dtype=str: 반/번호(예: 6/1)가 날짜로 자동 변환되는 것을 방지
        df = pd.read_excel(file_path, header=None, engine='openpyxl', dtype=str)
        
        # 2. 헤더 행 동적 탐색
        # 사용자가 언급한 '학년도', '학기' 또는 필수인 '과목', '성명' 등이 포함된 행을 찾습니다.
        header_row_idx = None
        for i, row in df.iterrows():
            row_str = " ".join(row.fillna('').astype(str).values)
            # 조건: '과목'이나 '과목명'이 있고, '성명'이나 '이름'이 있는 행을 헤더로 간주
            if ('과목' in row_str) and ('성명' in row_str or '이름' in row_str):
                header_row_idx = i
                break
        
        if header_row_idx is None:
            return "오류: 데이터 헤더(과목, 성명 등)를 찾을 수 없습니다. 엑셀 양식을 확인해주세요."

        # 3. 헤더 설정 및 데이터 재구조화
        df.columns = df.iloc[header_row_idx] # 찾은 행을 컬럼명으로 설정
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True) # 헤더 이후 데이터만 남김
        
        # 4. 컬럼명 정규화 (공백 제거)
        # 예: '과 목' -> '과목', '세부능력 및 특기사항' -> '세부능력및특기사항'
        df.columns = [str(c).replace(' ', '').replace('\n', '').strip() for c in df.columns]

        # 5. 필요한 컬럼 매핑 (유사한 이름 대응)
        col_map = {
            'subject': None,
            'name': None,
            'content': None,
            'grade': None,
            'class_info': None
        }

        for col in df.columns:
            if '과목' in col: col_map['subject'] = col
            elif '성명' in col or '이름' in col: col_map['name'] = col
            elif '세부능력' in col or '특기사항' in col: col_map['content'] = col
            elif '학년' in col and '학년도' not in col: col_map['grade'] = col # '학년도' 제외
            elif '반' in col and '번호' in col: col_map['class_info'] = col # '반/번호' 등
        
        # 필수 컬럼 체크
        if not col_map['content']:
            return "오류: '세부능력 및 특기사항' 컬럼을 찾을 수 없습니다."
        
        # 6. 데이터 전처리
        # (1) 필요한 컬럼만 추출
        valid_cols = [v for k, v in col_map.items() if v is not None]
        df = df[valid_cols].copy()

        # (2) 병합된 셀 처리 (Forward Fill) - '내용' 컬럼 제외
        # 메타데이터(학년, 반, 번호, 이름 등)만 앞의 값으로 채웁니다.
        content_col = col_map['content']
        name_col = col_map['name']
        meta_cols = [c for c in df.columns if c != content_col]
        
        df[meta_cols] = df[meta_cols].ffill()

        # (3) 내용 컬럼 처리
        # 내용이 없는 경우(NaN)를 빈 문자열로 변경 (삭제하지 않음!)
        df[content_col] = df[content_col].fillna("")

        # (4) 유효하지 않은 행 제거 (기준: 이름)
        # 내용이 없더라도 '이름'이 있다면 학생 데이터이므로 남겨야 합니다.
        # 이름조차 없다면 엑셀의 빈 하단부이거나 쓰레기 데이터일 확률이 높습니다.
        if name_col:
            df = df.dropna(subset=[name_col])
            # 이름이 빈 문자열인 경우도 제거
            df = df[df[name_col].astype(str).str.strip() != '']

        # 7. Markdown 생성
        md_output = []
        md_output.append("# 중학교 과목별 세부능력 및 특기사항 분석")
        md_output.append("---")
        
        for _, row in df.iterrows():
            content_raw = str(row[content_col]).strip()
            
            # [변경점] 내용이 비어있는 경우 명시적 표시 (선택 사항)
            # LLM에게 빈 칸임을 확실히 알리기 위해 마킹하거나, 그냥 빈 문자열을 줍니다.
            if content_raw == "":
                content_marked = "[내용 없음]" # 혹은 "" 그대로 두어도 됨
            else:
                content_marked = mark_multiple_spaces(content_raw)
            
            # 메타데이터 추출
            subject = row.get(col_map['subject'], '과목')
            name = row.get(col_map['name'], '이름')
            grade = row.get(col_map['grade'], '')
            class_info_raw = row.get(col_map['class_info'], '')
            
            # 반/번호 포맷팅
            class_info = str(class_info_raw)
            if '/' in class_info:
                parts = class_info.split('/')
                if len(parts) == 2:
                    class_info = f"{parts[0]}반 {parts[1]}번"
            
            grade_str = f"{grade}학년 " if grade else ""

            header = f"## [{subject}] {name} ({grade_str}{class_info})"
            md_output.append(header)
            md_output.append(f"> **기록 내용**\n\n{content_marked}")
            md_output.append("\n---\n")
            
        return "\n".join(md_output)

    except Exception as e:
        return f"변환 중 오류 발생: {e}\n(파일의 컬럼명에 '과목', '성명', '세부능력' 등이 포함되어 있는지 확인해주세요.)"

# --- 행동특성 및 종합의견 변환 함수 ---
def xlsx_behavior_to_markdown(file_path):
    """
    엑셀/CSV 파일을 읽어 '행동특성 및 종합의견'을 정리하여 Markdown으로 반환합니다.
    """
    import re  # [수정] 정규표현식 사용을 위해 추가
    
    # 1. 파일 확장자에 따른 로드 (CSV/Excel 호환)
    try:
        ext = os.path.splitext(file_path)[-1].lower()
        if ext == '.csv':
            # CSV일 경우 encoding 주의 (euc-kr 또는 utf-8)
            try:
                df = pd.read_csv(file_path, header=None, encoding='utf-8-sig')
            except:
                df = pd.read_csv(file_path, header=None, encoding='cp949')
        else:
            df = pd.read_excel(file_path, engine='openpyxl', header=None)
    except Exception as e:
        return f"파일 읽기 오류: {e}"

    # 2. 유효한 헤더 행 찾기
    header_row_idx = None
    for idx, row in df.iterrows():
        row_str = "".join(row.astype(str).values).replace(" ", "")
        if "행동특성" in row_str and ("성명" in row_str or "이름" in row_str):
            header_row_idx = idx
            break
            
    if header_row_idx is None:
        return "오류: '행동특성'과 '성명'이 포함된 헤더 행을 찾을 수 없습니다."

    # 3. 헤더 적용 및 데이터 슬라이싱
    df.columns = df.iloc[header_row_idx]
    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)

    # 4. 컬럼명 전처리
    df.columns = [str(col).replace(" ", "").strip() for col in df.columns]
    
    content_col = None
    for col in df.columns:
        if "행동특성" in col or "종합의견" in col:
            df.rename(columns={col: '행동특성및종합의견'}, inplace=True)
            content_col = '행동특성및종합의견'
            break
    
    if not content_col:
        return "오류: '행동특성 및 종합의견' 컬럼을 찾을 수 없습니다."

    # 공백 문자, 빈 문자열을 확실하게 NaN으로 변환
    df = df.replace(r'^\s*$', np.nan, regex=True)

    # 5. 데이터 정제 (Garbage 제거)
    keywords_to_remove = ['학교생활기록부', '사용자명', '담당교사']
    school_pattern = re.compile(r'^\s*.*(?:초등|중|고등)학교\s*$')

    def is_garbage(row):
        content = str(row.get(content_col, ''))
        name_val = str(row.get('성명', ''))
        grade_val = str(row.get('학년', ''))
        
        # 이름 칸에 불필요한 키워드가 있거나, 헤더가 반복되면 삭제
        if any(k in name_val for k in keywords_to_remove):
            return True
        if name_val.replace(" ", "") == "성명": 
            return True
        
        # [수정] 이름 칸에 학교명만 덩그러니 있는 경우 삭제
        if school_pattern.match(name_val):
            return True
            
        # 내용이나 학년 정보가 있으면 유효한 데이터로 간주
        has_meaningful_data = (
            content.strip() not in ('nan', '') and content is not None or
            grade_val.strip() not in ('nan', '') and grade_val is not None
        )
        
        return not has_meaningful_data

    df = df[~df.apply(is_garbage, axis=1)]

    # 6. 결측치 채우기 (Forward Fill)
    fill_cols = [c for c in ['학년', '반', '번호', '성명'] if c in df.columns]
    for col in fill_cols:
        df[col] = df[col].ffill()

    # 7. 데이터 병합 (Grouping)
    group_cols = [c for c in ['학년', '반', '번호', '성명'] if c in df.columns]
    
    if not group_cols:
        return "오류: 그룹화할 기준 컬럼(번호, 성명 등)이 없습니다."

    def join_text(text_series):
        # 중복된 텍스트 제거 및 병합
        texts = [str(t).strip() for t in text_series if str(t) != 'nan' and str(t).strip() != '']
        joined = " ".join(texts) # 공백 검사를 위해 \n 대신 한 칸 공백으로 합침
        return mark_multiple_spaces(joined) # 마킹 함수 적용

    df_grouped = df.groupby(group_cols)[content_col].apply(join_text).reset_index()

    # 8. 정렬
    sort_cols = [c for c in ['학년', '반', '번호'] if c in df_grouped.columns]
    if sort_cols:
        # 정렬을 위해 숫자형으로 변환 가능한 것은 변환
        for col in sort_cols:
            df_grouped[col] = pd.to_numeric(df_grouped[col])
        df_grouped = df_grouped.sort_values(by=sort_cols)

    # 9. Markdown 생성
    markdown_lines = []
    
    for _, row in df_grouped.iterrows():
        name = row.get('성명', '이름없음')
        number = row.get('번호', 0)
        grade = row.get('학년', '')
        content = row.get(content_col, '').strip()
        
        if not content: 
            continue

        # 번호 정수화
        try:
            number = int(number)
        except:
            pass

        # 학년 표시
        grade_str = f"{grade}학년" if str(grade).strip() else ""

        md_block = f"""
### 행동특성 및 종합의견 | {name} (No.{number})
- **학년:** {grade_str}
- **내용:**
> {content}

---
"""
        markdown_lines.append(md_block)

    return "".join(markdown_lines)

# 창의적 체험활동 (고등학교 2-3학년, 현재학년) xlsx 파일 처리
def xlsx_creative_activity_to_markdown(file_path):
    # 1. 파일 불러오기
    df = pd.read_excel(file_path, engine='openpyxl', header=None)
    
    # 2. 헤더 행 찾기
    header_row_idx = None
    for idx, row in df.iterrows():
        row_str = "".join(row.astype(str).values).replace(" ", "")
        if "특기사항" in row_str and "영역" in row_str:
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        return "오류: '영역'과 '특기사항'이 포함된 헤더 행을 찾을 수 없습니다."
    
    # 3. 데이터 슬라이싱 및 기초 정제
    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    
    # 불필요한 행 제거
    if df.shape[1] > 4:
        df = df[df.iloc[:, 4] != '영역']
    
    # '~학교'로 끝나는 텍스트가 있으면 해당 행 제거
    if df.shape[1] > 13:
        # (?: ... )는 비캡처 그룹, $는 문자열 끝을 의미
        df = df[~df.iloc[:, 13].astype(str).str.contains(r'(?:초등|중|고등)학교$', regex=True, na=False)]
        
    if df.shape[1] > 1:
        df = df[df.iloc[:, 1] != '번호']
    
    # 4. 열 선택 및 이름 변경
    try:
        target_columns_indices = [1, 2, 3, 4, 6, 8, 12]
        new_column_names = ['번호', '이름', '학년', '영역', '시간', '특기사항', '진로희망분야']
        
        df_clean = df.iloc[:, target_columns_indices].copy()
        df_clean.columns = new_column_names
    except IndexError:
        return "오류: 엑셀 파일의 열 개수가 맞지 않습니다. 양식을 확인해주세요."
    
    # [기초 작업] 번호와 이름 채우기
    df_clean[['번호', '이름', '학년']] = df_clean[['번호', '이름', '학년']].ffill()
    
    # [기초 작업] 진로희망분야 처리
    mask_career = (df_clean['영역'] == '진로활동') & (df_clean['특기사항'].astype(str).str.contains('희망분야'))
    df_clean.loc[mask_career, '특기사항'] = "희망분야: " + df_clean.loc[mask_career, '진로희망분야'].astype(str) + "\n"
    
    # ---------------------------------------------------------
    # [1차 병합] 물리적으로 끊긴 문장 이어붙이기 (시간이 NaN인 경우)
    # ---------------------------------------------------------
    df_clean['split_group_id'] = df_clean['시간'].notna().cumsum()
    df_merged_1 = df_clean.groupby(['번호', '이름', 'split_group_id'], as_index=False).agg({
        '학년': 'first',
        '영역': 'first',
        '시간': 'first',
        '특기사항': lambda x: mark_multiple_spaces("".join(x.dropna().astype(str)))
    })
    
    # ---------------------------------------------------------
    # [2차 병합] 연속된 동일 영역 합치기
    # ---------------------------------------------------------
    df_merged_1 = df_merged_1.sort_values(by=['번호', 'split_group_id'])
    
    condition = (df_merged_1['번호'] != df_merged_1['번호'].shift()) | \
                (df_merged_1['이름'] != df_merged_1['이름'].shift()) | \
                (df_merged_1['영역'] != df_merged_1['영역'].shift())
    df_merged_1['area_group_id'] = condition.cumsum()
    
    df_final = df_merged_1.groupby('area_group_id', as_index=False).agg({
        '번호': 'first',
        '이름': 'first',
        '학년': 'first',
        '영역': 'first',
        '시간': 'first',
        '특기사항': lambda x: mark_multiple_spaces("".join(x.dropna().astype(str)))
    })
    
    # [마무리 정제]
    df_final = df_final[df_final['영역'] != '창의적체험활동']
    df_final = df_final.drop(columns=['area_group_id'])
    
    # ---------------------------------------------------------
    # [Markdown 변환 로직]
    # ---------------------------------------------------------
    def generate_markdown(df):
        md_output = ""
        df_sorted = df.sort_values(by=['번호'])
        
        for (num, name, grade), group in df_sorted.groupby(['번호', '이름', '학년']):
            try:
                num_int = int(num)
                grade_int = int(grade)
            except:
                num_int = num
                grade_int = grade

            md_output += f"## {num_int}번 {name} ({grade_int}학년)\n\n"
            
            for _, row in group.iterrows():
                area = row['영역']
                time = row['시간']
                desc = row['특기사항']
                
                time_str = f"{int(time)}시간" if pd.notna(time) and str(time).replace('.','').isdigit() else "시간 미기재"
                
                md_output += f"### {area} ({time_str})\n"
                md_output += f"{desc}\n\n"
            
            md_output += "---\n\n"
            
        return md_output
    
    contents = generate_markdown(df_final)
    return contents

# 창의적 체험활동 (고등학교 2-3학년, 활동별)
def xlsx_creative_activity_to_markdown_sectional_high(input_xlsx):
    """
    엑셀 파일을 읽어 창의적 체험활동 데이터를 추출하고,
    공백 검출 함수를 적용한 뒤 Markdown Table 문자열로 반환합니다.
    """
    
    # 1. 파일 확장자에 따른 처리 방식 결정
    file_type = 'xlsx' if input_xlsx.endswith('.xlsx') else 'csv'
    rows = []

    # 2. 파일 읽기
    if file_type == 'xlsx':
        wb = openpyxl.load_workbook(input_xlsx, data_only=True)
        # sheet1 혹은 활성 시트 선택
        if 'sheet1' in [s.lower() for s in wb.sheetnames]:
            sheet_name = [s for s in wb.sheetnames if s.lower() == 'sheet1'][0]
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # openpyxl로 데이터 읽기
        rows = list(ws.iter_rows(values_only=True))

    elif file_type == 'csv':
        with open(input_xlsx, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = [ [ (c if c.strip() != '' else None) for c in r ] for r in reader ]

    # 3. 데이터 처리
    processed_data = []
    current_record = None
    
    # 엑셀 열 인덱스 (0부터 시작)
    IDX_NO = 1       # B
    IDX_NAME = 2     # C
    IDX_GRADE = 4    # E
    IDX_AREA = 5     # F
    IDX_TIME = 7     # H
    IDX_SPEC = 9     # J
    IDX_HOPE = 13    # N

    for row in rows:
        if len(row) <= IDX_SPEC:
            continue
            
        val_no = row[IDX_NO]
        val_name = row[IDX_NAME]
        val_spec = row[IDX_SPEC]
        
        # 문자열 변환
        str_no = str(val_no).strip() if val_no else ""
        str_name = str(val_name).strip() if val_name else ""
        
        # 헤더/푸터 필터링
        is_header = "번 호" in str_no or "성  명" in str_name or "창의적 체험활동상황" in str_name
        # '초등학교', '중학교', '고등학교'가 포함되어 있는지 정규식으로 확인
        is_school_name = re.search(r"(초등학교|중학교|고등학교)", str_name)
        is_footer = is_school_name or (str_no.isdigit() and "/" in str_name)

        if is_header or is_footer:
            continue
        
        # [새 레코드]
        if val_no is not None and str_no.isdigit():
            if current_record:
                processed_data.append(current_record)
            
            current_record = {
                '번호': val_no,
                '성명': val_name,
                '학년': row[IDX_GRADE],
                '영역': row[IDX_AREA],
                '시간': row[IDX_TIME],
                '특기사항_parts': [val_spec] if val_spec else [],
                '희망분야': row[IDX_HOPE] if len(row) > IDX_HOPE else None
            }
        
        # [내용 이어짐]
        elif current_record is not None:
            if val_spec:
                current_record['특기사항_parts'].append(val_spec)
            
            if len(row) > IDX_HOPE and row[IDX_HOPE]:
                if not current_record['희망분야']:
                    current_record['희망분야'] = row[IDX_HOPE]

    if current_record:
        processed_data.append(current_record)

    # 4. 최종 데이터 구성
    final_list = []
    for item in processed_data:
        # 페이지 넘김 병합 (공백 1칸)
        parts = [str(p).strip() for p in item['특기사항_parts'] if p]
        full_specialty = "".join(parts)
        
        hope_field = str(item['희망분야']).strip() if item['희망분야'] else ""
        
        # 진로활동 희망분야 처리
        if item['영역'] and "진로활동" in str(item['영역']):
            full_specialty = re.sub(r'^희망분야[:]?\s*', '', full_specialty)
            
            if hope_field:
                full_specialty = f"희망분야: {hope_field}\n{full_specialty}"
        
        # [공백 검출 함수 적용] 최종 텍스트에 대해 공백 검사 수행
        full_specialty = mark_multiple_spaces(full_specialty)

        final_list.append({
            '번호': item['번호'],
            '성명': item['성명'],
            '학년': item['학년'],
            '영역': item['영역'],
            '시간': item['시간'],
            '특기사항': full_specialty
        })

    # 5. DataFrame 생성 및 Markdown 변환
    df = pd.DataFrame(final_list)
    
    if df.empty:
        return ""
    
    # tabulate로 Markdown 변환
    markdown_output = tabulate(df, headers='keys', tablefmt='github', showindex=False)
    
    return markdown_output

# 창의적 체험활동 (중학교, 활동별)
def xlsx_creative_activity_to_markdown_sectional_mid(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        extracted_data = []
        
        current_info = {
            "번호": None,
            "성명": None,
            "학년": None,
            "영역": None,
            "시간": None,
            "특기사항": ""
        }
        
        IDX_NO = 1
        IDX_NAME = 2
        IDX_GRADE = 4
        IDX_AREA = 5
        IDX_TIME = 7
        IDX_NOTE = 9 
        
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(cell).strip() if cell is not None else "" for cell in row]
            
            if len(row_vals) <= IDX_NOTE:
                continue
        
            val_no = row_vals[IDX_NO]
            val_name = row_vals[IDX_NAME]
            val_area = row_vals[IDX_AREA]
            val_note = row_vals[IDX_NOTE]
        
            if "번 호" in val_no or "성  명" in val_name:
                continue

            if val_note == "특기사항":
                continue
            
            if val_area == "영역" or "시 간" in str(row_vals[IDX_TIME]):
                continue

            if val_no.isdigit() and val_name != "":
                if current_info["성명"]:
                    extracted_data.append(current_info)
                
                initial_note = mark_multiple_spaces(val_note)
                
                if val_area == "진로활동" and val_note == "희망분야":
                    hope_field = ""
                    for cell_val in row_vals[IDX_NOTE+1:]:
                        if cell_val:
                            hope_field = cell_val
                            break
                    
                    if hope_field:
                        initial_note = f"희망분야: {hope_field}\n"
                    else:
                        initial_note = "" 
                
                current_info = {
                    "번호": val_no,
                    "성명": val_name,
                    "학년": row_vals[IDX_GRADE],
                    "영역": val_area,
                    "시간": row_vals[IDX_TIME],
                    "특기사항": initial_note 
                }
            
            elif val_no == "" and val_name == "" and val_note != "":
                if val_note == "특기사항": 
                    continue

                if current_info["성명"]:
                    # 띄어쓰기 한 칸과 함께 마킹 적용
                    current_info["특기사항"] += " " + mark_multiple_spaces(val_note)
        
        if current_info["성명"]:
            extracted_data.append(current_info)
        
        df_final = pd.DataFrame(extracted_data)
        
        if df_final.empty:
            return "오류: 추출된 데이터가 없습니다. 파일 양식을 확인해주세요."

        df_final["번호"] = pd.to_numeric(df_final["번호"], errors='coerce')

        markdown_output = []
        
        for index, row in df_final.iterrows():
            section_header = f"### [{row['영역']}] {row['번호']}번 {row['성명']} ({row['학년']}학년)"
            
            content = (
                f"{section_header}\n"
                f"- **이수시간**: {row['시간']}시간\n"
                f"- **특기사항**:\n"
                f"{row['특기사항']}\n"
                f"\n---\n"
            )
            markdown_output.append(content)
        
        return "".join(markdown_output)

    except Exception as e:
        return f"오류: 창의적 체험활동(중학교, 활동별) 처리 중 문제가 발생했습니다. {str(e)}"

# 창의적 체험활동 (고등학교 1학년, 현재학년)
def xlsx_creative_activity_to_markdown_1st_year(file_path):
    # 1. 엑셀 파일 로드
    df = pd.read_excel(file_path, sheet_name='sheet1', engine='openpyxl', header=None)
    
    refined_list = []
    current_student = {"no": None, "name": None, "grade": None}
    
    target_categories = ['자율·자치활동', '동아리활동', '진로활동']
    
    for i in range(len(df)):
        row = df.iloc[i]
        
        # 2. 학생 정보 업데이트 (번호가 있는 행에서만 갱신)
        raw_no = str(row[1]).strip() if pd.notna(row[1]) else ""
        if raw_no.isdigit():
            current_student["no"] = raw_no
            current_student["name"] = str(row[2]).strip()
            current_student["grade"] = str(row[3]).strip()
        
        # 3. 영역 확인 (E열)
        category = str(row[4]).strip() if pd.notna(row[4]) else ""
        
        # 4. 특기사항 내용(I열) 정제 및 추출
        main_content = str(row[8]).strip() if pd.notna(row[8]) and str(row[8]).lower() != 'nan' else ""
        clean_content = re.sub(r'^(진로희망|희망분야)\s*[:：]?\s*', '', main_content)
        
        # 다중 공백 마킹 적용
        clean_content = mark_multiple_spaces(clean_content)
        
        if clean_content in ["특기사항", "내용", "진로희망", "희망분야"]:
            clean_content = ""

        # 5. 데이터 처리 (영역이 지정된 경우)
        if category in target_categories:
            # [이 부분이 누락되어 오류가 발생했습니다]
            hours = str(row[6]).strip() if pd.notna(row[6]) else "0"
            hope_field = str(row[12]).strip() if category == '진로활동' and pd.notna(row[12]) else ""
            
            # 페이지 넘김으로 인한 중복 데이터 확인 (이전 항목과 학생 & 영역이 같은 경우)
            if (refined_list and 
                refined_list[-1]['번호'] == current_student["no"] and 
                refined_list[-1]['영역'] == category):
                
                if clean_content:
                    prev_content = refined_list[-1]['내용뭉치']
                    refined_list[-1]['내용뭉치'] = (prev_content + "" + clean_content).strip()
            else:
                # 새로운 항목 추가
                refined_list.append({
                    '번호': current_student["no"],
                    '성명': current_student["name"],
                    '학년': current_student["grade"],
                    '영역': category,
                    '시간': hours, # 이제 정상적으로 정의됨
                    '희망분야': hope_field, # 이제 정상적으로 정의됨
                    '내용뭉치': clean_content
                })
        
        # 6. 영역명은 없지만 내용만 이어지는 행 처리 (I열 내용 존재 시)
        elif category == "" and current_student["no"] is not None and refined_list:
            if clean_content:
                prev_content = refined_list[-1]['내용뭉치']
                refined_list[-1]['내용뭉치'] = (prev_content + "" + clean_content).strip()

    # 7. 최종 텍스트 조립 (희망분야 및 특기사항 결합)
    final_processed = []
    for item in refined_list:
        content_str = item['내용뭉치'].strip()
        
        if item['영역'] == '진로활동':
            # 희망분야 문구 정제
            clean_hope = re.sub(r'^(진로희망|희망분야)\s*[:：]?\s*', '', item['희망분야']).strip()
            if clean_hope:
                final_note = f"희망분야: {clean_hope} | {content_str}"
            else:
                final_note = content_str
        else:
            final_note = content_str
            
        final_processed.append({
            '번호': item['번호'],
            '성명': item['성명'],
            '학년': item['학년'],
            '영역': item['영역'],
            '시간': item['시간'],
            '특기사항': final_note
        })
    
    # 8. 데이터프레임 생성 및 마크다운 변환
    df_refined = pd.DataFrame(final_processed)
    contents = tabulate(df_refined, headers='keys', tablefmt='pipe', showindex=False)
    
    return contents

# 창의적 체험활동 (1학년, 활동별)
def xlsx_creative_activity_to_markdown_1st_year_sectional(file_path):

    # 창의적 체험활동 (1학년, 활동별) markdown으로 바꾸는 함수
    def df_to_llm_markdown(df):
        """
        DataFrame을 LLM이 이해하기 쉬운 Markdown 형식의 텍스트로 변환합니다.
        """
        if df.empty:
            return "데이터가 없습니다."
    
        markdown_lines = []
        
        for index, row in df.iterrows():
            # 각 행(row)을 하나의 섹션으로 구성
            section = (
                f"## 학생 정보: {row['번호']} {row['성명']}\n"
                f"- **학년**: {row['학년']}\n"
                f"- **영역**: {row['영역']}\n"
                f"- **이수 시간**: {row['시간']}\n"
                f"- **특기사항**:\n"
                f"{row['특기사항']}\n"
            )
            markdown_lines.append(section)
        
        # 각 항목을 구분선으로 나눔
        return "\n---\n".join(markdown_lines)

    
    if not os.path.exists(file_path):
        return "Error: 파일을 찾을 수 없습니다."

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'sheet1' in wb.sheetnames:
        ws = wb['sheet1']
    else:
        ws = wb.active

    data = []
    last_entry = None
    start_processing = False
    
    footer_pattern = re.compile(r'.*(초|중|고)등학교|.*학교생활기록부')

    for row in ws.iter_rows(values_only=True):
        # 인덱스: B(1), C(2), E(4), F(5), H(7), J(9), N(13)
        val_b = row[1]  # 번호
        val_c = row[2]  # 성명
        val_e = row[4]  # 학년
        val_f = row[5]  # 영역
        val_h = row[7]  # 시간
        val_j = row[9]  # 특기사항
        val_n = row[13] # 희망분야

        def clean_str(s):
            return str(s).strip() if s is not None else ""

        str_b = clean_str(val_b)
        str_c = clean_str(val_c)
        str_f = clean_str(val_f)
        str_j = clean_str(val_j)
        str_n = clean_str(val_n)

        # 1. 헤더 감지
        if "번 호" in str_b or "성  명" in str_c:
            start_processing = True
            continue
        
        if not start_processing:
            continue
            
        # 2. [수정] 바닥글/페이지 구분자 건너뛰기 (범용 로직 적용)
        # B열(번호) 또는 J열(특기사항)에 학교명 패턴이나 '학교생활기록부'가 포함되면 건너뜁니다.
        if str_j and footer_pattern.search(str_j):
            continue
        if str_b and footer_pattern.search(str_b):
            continue

        # 3. 데이터 처리 로직 판단
        is_new_student = False
        
        # 번호가 있는 경우
        if str_b and (str_b.isdigit() or len(str_b) < 5):
            # 이전 학생과 번호/성명이 같으면 페이지 넘어감에 의한 중복 -> 이어지는 내용으로 처리
            if last_entry and last_entry['번호'] == str_b and last_entry['성명'] == str_c:
                is_new_student = False
            else:
                is_new_student = True
        else:
            is_new_student = False

        # --- 로직 실행 ---
        
        if is_new_student:
            # [새로운 항목 등록]
            note_content = str_j
            
            # 진로활동 특수 처리
            if "진로" in str_f and str_n:
                # J열 내용이 '희망분야'라는 단어로 시작하면 제거 (중복 방지)
                note_content = re.sub(r'^희망분야\s*', '', note_content)
                
                # 포맷팅 적용
                formatted_note = f"희망분야: {str_n}\n{note_content}"
            else:
                formatted_note = note_content
            
            entry = {
                '번호': str_b,
                '성명': str_c,
                '학년': clean_str(val_e),
                '영역': str_f,
                '시간': clean_str(val_h),
                '특기사항': formatted_note
            }
            data.append(entry)
            last_entry = entry
            
        elif str_j:
            # [내용 이어 붙이기]
            # 헤더 반복 제외
            if "특기사항" in str_j:
                continue
                
            if last_entry:
                last_entry['특기사항'] += "" + str_j

    # DataFrame 생성
    df = pd.DataFrame(data)
    
    # 공백 검출 함수 적용 (외부 함수로 가정하고 예외 처리 추가)
    if not df.empty and '특기사항' in df.columns:
        try:
            # mark_multiple_spaces 함수가 정의되어 있다고 가정
            df['특기사항'] = df['특기사항'].apply(mark_multiple_spaces)
        except NameError:
            pass # 해당 함수가 없으면 패스

    # Markdown으로 변환하여 반환
    markdown_text = df_to_llm_markdown(df)
    
    return markdown_text

# 동아리 활동
def xlsx_club_to_markdown(file_path):
    """
    동아리 활동 엑셀 파일을 읽어 Markdown Table 형식으로 반환
    """
    try:
        # 1. 엑셀 파일 읽기 (헤더 없이 일단 읽음)
        # sheet_name=0 : 시트 이름이 달라도 첫 번째 시트를 읽음
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None, engine='openpyxl')
        
        # 2. 동적 헤더 찾기 ('성명', '특기사항' 등이 포함된 행을 헤더로 인식)
        header_row_idx = None
        for idx, row in df_raw.iterrows():
            row_str = "".join(row.astype(str).values).replace(" ", "")
            if "성명" in row_str and "특기사항" in row_str and "반" in row_str:
                header_row_idx = idx
                break
        
        if header_row_idx is None:
            # 헤더를 못 찾은 경우 기본값(3행) 시도 혹은 에러 반환
            # 여기서는 안전하게 3행(0부터 시작하므로 index 3)을 가정하되 경고
            header_row_idx = 3

        # 3. 헤더 적용 및 데이터 슬라이싱
        df_raw.columns = df_raw.iloc[header_row_idx]
        df = df_raw.iloc[header_row_idx + 1:].reset_index(drop=True)
        
        # 4. 컬럼명 공백 제거 (예: '학 년' -> '학년')
        df.columns = [str(c).replace(' ', '').strip() for c in df.columns]

        # 5. 필요한 컬럼만 추출
        target_cols = ['학년', '반', '번호', '성명', '이수시간', '특기사항']
        
        # 필수 컬럼이 다 있는지 확인
        missing_cols = [col for col in target_cols if col not in df.columns]
        if missing_cols:
             # 만약 '이수시간' 컬럼명이 다를 경우(예: 시간) 처리
             if '시간' in df.columns and '이수시간' in missing_cols:
                 df.rename(columns={'시간': '이수시간'}, inplace=True)
             else:
                 return f"오류: 엑셀 파일에서 다음 열을 찾을 수 없습니다: {missing_cols}"
        
        df = df[target_cols].copy()

        # 6. 전처리: 데이터 정제
        # 학년, 반 등이 비어있는 병합 셀 채우기
        df[['학년', '반', '번호']] = df[['학년', '반', '번호']].ffill()

        # '성명'이 있는 행을 기준으로 그룹 ID 생성
        df['group_id'] = df['성명'].notna().cumsum()

        # 7. 데이터 병합 (핵심 로직)
        final_df = df.groupby('group_id').agg({
            '학년': 'first',
            '반': 'first',
            '번호': 'first',
            '성명': 'first',
            '이수시간': 'first',
            '특기사항': lambda x: mark_multiple_spaces("".join(x.dropna().astype(str)))
        }).reset_index(drop=True)
        
        # 8. 불필요한 행(헤더 반복 등) 제거
        # 성명이 '성명'인 행이나 숫자가 아닌 학년 제거
        final_df = final_df[final_df['성명'] != '성명']

        # 9. 보기 좋게 형변환
        cols_to_int = ['학년', '반', '번호', '이수시간']
        for col in cols_to_int:
            # 숫자가 아닌 문자가 섞여 있을 경우를 대비해 coerce 사용 후 0으로 채움
            final_df[col] = pd.to_numeric(final_df[col], errors='coerce').fillna(0).astype(int)

        # 10. Markdown 변환 (Pandas 내장 기능 사용 - tabulate 의존성 문제 해결)
        # index=False로 하여 불필요한 인덱스 번호 제거
        contents = final_df.to_markdown(index=False)
        
        return contents

    except Exception as e:
        return f"오류: 동아리 파일 처리 중 문제가 발생했습니다. {str(e)}"

#-----------------------------------------------------------------------------------------------------------------------#
# --- Chainlit 앱 로직 ---

async def send_subject_selection():
    actions = [
        cl.Action(name="subject_select", value=subject, label=subject, payload={"subject": subject})
        for subject in PROMPTS.keys()
    ]
    
    await cl.Message(
        content="검토할 영역을 선택해주세요.",
        actions=actions,
        author="생기부 검토 도우미"
    ).send()

@cl.on_chat_start
async def on_chat_start():
    """앱 시작 시 API Key 입력 요구 및 검증 후 메뉴 표시"""
    
    # 1. API Key 입력 루프
    while True:
        res = await cl.AskUserMessage(
            content="안녕하세요. 학교생활기록부 검토 도우미입니다.\n"
                    "**Google Gemini API 키**를 입력해주세요.\n\n"
                    "API 키와 생기부 내용 및 검토 결과는 서버에 전혀 기록하지 않습니다. 안심하고 사용하세요..\n"
                    "무료 API 키로는 **하루 최대 20회**까지 NEIS 엑셀 파일을 통째로 올려서 검토할 수 있습니다.\n"
                    "Google Gemini API 키는 https://aistudio.google.com/app/api-keys 에서 무료로 발급받을 수 있습니다.", 
            timeout=600
        ).send()

        if res:
            user_api_key = res["output"].strip()
            
            # 검증 중 메시지
            msg = cl.Message(content=f"**{TEST_MODEL_NAME}** 모델로 API 키 연결 확인 중...", author="System")
            await msg.send()
            
            # Lite 모델로 빠르게 검증
            is_valid = await validate_api_key(user_api_key)
            
            if is_valid:
                # 유효하면 세션에 저장하고 루프 탈출
                cl.user_session.set("user_api_key", user_api_key)
                msg.content = "✅ API Key가 확인되었습니다."
                await msg.update()
                break
            else:
                # [수정] 실패 메시지로 업데이트
                msg.content = "❌ 유효하지 않은 API Key입니다. 다시 입력해주세요."
                await msg.update()
        else:
            # 타임아웃
            await cl.Message(content="입력 시간이 초과되었습니다. 페이지를 새로고침 해주세요.").send()
            return

    # 2. 인사말 및 영역 선택
    await cl.Message(content=f"분석 모델: {MAIN_MODEL_NAME}", author="생기부 검토 도우미").send()
    await send_subject_selection()

@cl.action_callback("subject_select")
async def on_subject_select(action: cl.Action):
    """영역 선택 시 세션에 저장하고 안내를 표시합니다."""
    subject = action.payload["subject"]
    cl.user_session.set("subject", subject)
    cl.user_session.set("session_id", cl.user_session.get("id"))
    
    # 선택 확인 메시지
    await cl.Message(content=f"**{subject}** 영역을 선택했습니다.", author="생기부 검토 도우미").send()
    
    guide_message = ""

    if subject == "과목별 세부능력 및 특기사항 (생기부 영역별 출력)":
        guide_message = (
            "1. NEIS 학급담임-학생부-학교생활기록부-학생부 항목별 조회 출력 메뉴에서 과세특(활동별)을 조회하여 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "과세특 (교사 입력 과세특 엑셀 파일, 고등학교)":
        guide_message = (
            "1. NEIS 교과담임-성적-성적처리에서 과목별세부능력및특기사항을 조회하여 [엑셀내려받기] 버튼을 눌러 엑셀 파일(.xlsx)을 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "과세특 (교사 입력 과세특 엑셀 파일, 중학교)":
        guide_message = (
            "1. NEIS 교과담임-성적-성적처리에서 과목별세부능력및특기사항을 조회하여 [엑셀내려받기] 버튼을 눌러 엑셀 파일(.xlsx)을 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "행동특성 및 종합의견":
        guide_message = (
            "1. NEIS 학급담임-학생부-학교생활기록부-학생부 항목별 조회 출력 메뉴에서 행동특성 및 종합의견을 조회하여 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "창의적 체험활동 (고등학교 2-3학년, 현재학년)":
        guide_message = (
            "1. NEIS 학급담임-학생부-학교생활기록부-학생부 항목별 조회 출력 메뉴에서 창의적 체험활동 (현재학년)을 조회하여 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "창의적 체험활동 (고등학교 2-3학년, 활동별)":
        guide_message = (
            "1. NEIS 학급담임-학생부-학교생활기록부-학생부 항목별 조회 출력 메뉴에서 창의적 체험활동 (활동별)을 조회하여 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "창의적 체험활동 (중학교, 활동별)":
        guide_message = (
            "1. NEIS 학급담임-학생부-학교생활기록부-학생부 항목별 조회 출력 메뉴에서 창의적 체험활동 (활동별)을 조회하여 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "창의적 체험활동 (고등학교 1학년, 현재학년)":
        guide_message = (
            "1. NEIS 학급담임-학생부-학교생활기록부-학생부 항목별 조회 출력 메뉴에서 창의적 체험활동 (현재학년)을 조회하여 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "창의적 체험활동 (고등학교 1학년, 활동별)":
        guide_message = (
            "1. NEIS 학급담임-학생부-학교생활기록부-학생부 항목별 조회 출력 메뉴에서 창의적 체험활동 (활동별)을 조회하여 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    elif subject == "동아리 (교사 입력 엑셀 데이터 파일)":
        guide_message = (
            "1. NEIS 동아리담임-창의적체험활동-동아리활동관리-학생부자료기록-조회-출력-**XLS Data** 버튼을 눌러서 엑셀 파일로 저장하세요.\n"
            "2. 아래에 있는 파일 첨부 버튼을 눌러 파일(.xlsx)을 업로드해주세요.\n\n"
            "[참고1] NEIS에서 xlsx 파일을 받아서 Microsoft Excel로 열면 파일이 손상되었다고 나오는 경우가 있습니다. 이때는 복구하신 후, 그 파일을 올리면 정상적으로 인식이 될 것입니다.\n"
            "[참고2] 요즘 구글 Gemini 사용량이 많은지 구글에서 503 Service Unavailable 오류가 많이 납니다. 503 오류가 나면 구글 서버 잘못이니 당황하지 마시고, 몇 분 있다가 다시 시도하세요."
        )
    else:
        guide_message = "엑셀 파일(.xlsx)을 업로드해주세요."

    # 분기 처리된 메시지 전송
    await cl.Message(
        content=guide_message,
        author="생기부 검토 도우미"
    ).send()
    await action.remove()

@cl.on_message
async def on_message(message: cl.Message):
    subject = cl.user_session.get("subject")
    session_id = cl.user_session.get("session_id") or cl.user_session.get("id")
    
    # [추가] 세션에서 사용자 API Key 가져오기
    user_api_key = cl.user_session.get("user_api_key")
    
    # API Key가 없으면 중단 (새로고침 유도)
    if not user_api_key:
        await cl.Message(
            content="⚠️ API Key가 만료되었거나 없습니다. 페이지를 새로고침(F5)하여 키를 다시 입력해주세요.",
            author="System"
        ).send()
        return
    
    # 세션 ID가 없는 경우 방어 로직
    if not session_id:
        session_id = cl.user_session.get("id")

    if not subject:
        await cl.Message(
            content="먼저 검토할 영역을 선택해주세요.",
            author="생기부 검토 도우미"
        ).send()
        return
    
    files = [f for f in message.elements if isinstance(f, cl.File)]
    
    if not files:
        await cl.Message(
            content="xlsx 파일을 업로드해주세요.",
            author="생기부 검토 도우미"
        ).send()
        return
    
    uploaded_file = files[0]
    
    if not uploaded_file.name.lower().endswith((".xlsx", ".xls")):
        error_msg = "잘못된 파일 형식 업로드 (.xlsx 아님)"
        await cl.Message(
            content="xlsx 또는 xls 파일만 업로드할 수 있습니다.",
            author="생기부 검토 도우미"
        ).send()
        await cl.make_async(log_to_db)(session_id, subject, False, error_msg)
        return
    
    try:
        await cl.Message(
            content=f"'{uploaded_file.name}' 파일의 검토를 시작합니다.\n\n텍스트를 처리합니다. LLM의 기술적 한계를 감안하여 검토 의견을 읽어 주세요.\n\n답변이 나오기까지는 최대 3분 정도 걸립니다.",
            author="생기부 검토 도우미"
        ).send()
        
        # 파일 변환 로직 (기존과 동일하게 함수 호출)
        # (주의: 이전에 정의한 함수명을 정확히 사용해야 합니다)
        if subject == "과목별 세부능력 및 특기사항 (생기부 영역별 출력)":
            extracted_text = await cl.make_async(excel_to_clean_markdown_subject)(uploaded_file.path)
        elif subject == "과세특 (교사 입력 엑셀 파일, 고등학교)":
            extracted_text = await cl.make_async(xlsx_subject_to_markdown_high)(uploaded_file.path)
        elif subject == "과세특 (교사 입력 엑셀 파일, 중학교)":
            extracted_text = await cl.make_async(xlsx_subject_to_markdown_mid)(uploaded_file.path)
        elif subject == "행동특성 및 종합의견":
            extracted_text = await cl.make_async(xlsx_behavior_to_markdown)(uploaded_file.path)
        elif subject == "창의적 체험활동 (고등학교 2-3학년, 현재학년)":
            extracted_text = await cl.make_async(xlsx_creative_activity_to_markdown)(uploaded_file.path)
        elif subject == "창의적 체험활동 (고등학교 2-3학년, 활동별)":
            extracted_text = await cl.make_async(xlsx_creative_activity_to_markdown_sectional_high)(uploaded_file.path)
        elif subject == "창의적 체험활동 (중학교, 활동별)":
            extracted_text = await cl.make_async(xlsx_creative_activity_to_markdown_sectional_mid)(uploaded_file.path)
        elif subject == "창의적 체험활동 (고등학교 1학년, 현재학년)":
            extracted_text = await cl.make_async(xlsx_creative_activity_to_markdown_1st_year)(uploaded_file.path)
        elif subject == "창의적 체험활동 (고등학교 1학년, 활동별)":
            extracted_text = await cl.make_async(xlsx_creative_activity_to_markdown_1st_year_sectional)(uploaded_file.path)
        elif subject == "동아리 (교사 입력 엑셀 파일)":
            extracted_text = await cl.make_async(xlsx_club_to_markdown)(uploaded_file.path)
        else:
            extracted_text = "선택된 영역에 대한 처리 함수가 없습니다."

        if extracted_text.startswith("오류:"):
            await cl.Message(
                content=f"파일 처리 중 문제가 발생했습니다.\n\n{extracted_text}",
                author="생기부 검토 도우미"
            ).send()
            await cl.make_async(log_to_db)(session_id, subject, False, extracted_text)
            return
        
        system_role = PROMPTS[subject]
        
        messages = [
            SystemMessage(content=system_role),
            HumanMessage(content=f"생활기록부 본문 내용:\n\n{extracted_text}")
        ]
        
        msg = cl.Message(content="", author="검토 의견")
        await msg.send()
        
        full_response = []

        # 사용자 키로 메인 모델(gemini-3-flash-preview) 동적 생성
        dynamic_llm = ChatGoogleGenerativeAI(model=MAIN_MODEL_NAME, temperature=1.0, google_api_key=user_api_key)
        
        async with cl.Step(name="생활기록부 검토 중..."):
            async for chunk in dynamic_llm.astream(messages):
                if isinstance(chunk.content, list):
                    # 리스트인 경우 텍스트만 추출
                    content = ''.join([
                        item.get('text', '') if isinstance(item, dict) else str(item)
                        for item in chunk.content
                    ])
                elif isinstance(chunk.content, str):
                    content = chunk.content
                else:
                    # 예상치 못한 타입은 문자열로 변환
                    content = str(chunk.content)
                
                if content:  # 빈 문자열이 아닐 때만 처리
                    await msg.stream_token(content)
                    full_response.append(content)
            
            await msg.update()
        
        # 성공 로그
        await cl.make_async(log_to_db)(session_id, subject, True, None)
        
        cl.user_session.set("subject", None)
        
        await cl.Message(
            content="✅ 검토가 완료되었습니다. 계속하시려면 아래에서 영역을 다시 선택해주세요.",
            author="생기부 검토 도우미"
        ).send()
        
        await send_subject_selection()
    
    # [수정] 503 Service Unavailable 및 기타 구글 API 에러 구체적 처리
    except ServiceUnavailable as e:
        error_user_msg = "⚠️ 현재 AI 모델 사용량이 폭주하여 일시적으로 응답할 수 없습니다 (503 Error). 잠시 후(약 1분 뒤) 다시 시도해 주세요."
        await cl.Message(content=error_user_msg, author="시스템 오류").send()
        # DB에는 원본 에러 메시지 기록
        await cl.make_async(log_to_db)(session_id, subject, False, f"503 Service Unavailable: {str(e)}")

    except InternalServerError as e:
        error_user_msg = "⚠️ 구글 AI 서버 내부 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."
        await cl.Message(content=error_user_msg, author="시스템 오류").send()
        await cl.make_async(log_to_db)(session_id, subject, False, f"500 Internal Error: {str(e)}")

    except google_exceptions.InvalidArgument as e:
        error_msg = f"LLM Input Error: {str(e)}"
        if "request is too large" in str(e) or "token" in str(e).lower():
            await cl.Message(
                content="오류: 입력 내용이 너무 깁니다. 파일 내용을 줄여서 다시 시도해주세요.",
                author="오류"
            ).send()
        else:
            await cl.Message(content=f"API 요청 오류: {e}", author="오류").send()
        await cl.make_async(log_to_db)(session_id, subject, False, error_msg)

    except Exception as e:
        error_msg = str(e)
        if "API_KEY" in error_msg or "403" in error_msg:
             await cl.Message(
                content="⛔ **API Key 오류**: 키가 올바르지 않거나 허용량이 초과되었습니다. 새로고침 후 다시 시도해주세요.",
                author="오류"
            ).send()
        else:
             # 기존 에러 처리 로직 (503, 500 등)으로 연결하거나 직접 메시지 전송
             await cl.Message(content=f"오류 발생: {e}", author="오류").send()
        
        await cl.make_async(log_to_db)(session_id, subject, False, error_msg)
    
    except Exception as e:
        # 그 외 예상치 못한 모든 에러
        error_msg = f"Unexpected Error: {str(e)}"
        await cl.Message(
            content=f"처리 중 예기치 않은 오류가 발생했습니다: {e}",
            author="생기부 검토 도우미"
        ).send()
        await cl.make_async(log_to_db)(session_id, subject, False, error_msg)